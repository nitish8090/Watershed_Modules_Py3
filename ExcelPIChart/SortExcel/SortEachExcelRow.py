import pandas as pd
import numpy as np
import openpyxl


class SubBasin:
    def __init__(self, name, area, values):
        self.name = name
        self.area = area
        self.values = values

    def __str__(self):
        return "{}, {}, {}".format(self.name, self.area, self.values)


def main():
    path = r"book3CIMP5 RCP 4.5.xlsx"
    new_excel = r"sorted3_book3CIMP5 RCP 4.5.xlsx"
    writer = pd.ExcelWriter(new_excel, engine='openpyxl')

    wb = openpyxl.load_workbook(path)
    sheets = wb.sheetnames
    print(sheets)

    for sheet in sheets:
        print (sheet)
        df = pd.read_excel(path, sheet_name=sheet)
        df = df.iloc[:, 0:7]
        columns = (list(df.columns))

        new_df_list = []
        for i, row in df.iterrows():
            # print(row[0])
            subbasin = SubBasin(name=row[0], area=row[1], values=list(row[2:]))
            subbasin.values.sort()
            if not sheet.startswith("Vul"):
                subbasin.values.reverse()
                # print("Reverse Triggered")
            # print(subbasin)

            new_list = [subbasin.name, subbasin.area, *subbasin.values]
            # print(new_list)
            new_df_list.append(new_list)

        new_df = pd.DataFrame(new_df_list, columns=columns)
        # print(new_df)

        # book = openpyxl.load_workbook(new_excel)

        new_df.to_excel(writer, sheet_name=sheet, index=False)

        # break
    writer.save()
    writer.close()


main()
