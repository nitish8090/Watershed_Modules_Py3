import openpyxl
import pandas as pd
import os
import xlwings as xw

def main():
    project_file_path = r"F:\Watershed_Works\2102 Feb 2021\4_Anicut_Report\Process\WHS_Anicut_mod.xlsx"
    proposed_points = r"F:\Watershed_Works\2102 Feb 2021\4_Anicut_Report\Round2\Input\SelectedPoints.csv"
    output_folder = r"F:\Watershed_Works\2102 Feb 2021\4_Anicut_Report\Round2\Output\Excels"

    df = pd.read_csv(proposed_points, index_col=0)
    print(df)
    print(df.columns)

    wb = xw.Book(project_file_path)
    sht = wb.sheets[0]
    # print(sht.range('E2').value)

    # sht.range('E2').value = 86
    # wb.save(output_folder + "\\tes.xlsx")

    # exit()
    # book = openpyxl.load_workbook(project_file_path)
    # sheets = book.sheetnames
    # ws = book[sheets[0]]
    #
    for i in df.index:
        print("Working for ID: {}".format(i))
        District = df.loc[i]['District']
        latitude = df.loc[i]['Latitude']
        longitude = df.loc[i]['Longitude']
        catchment = df.loc[i]['Catchment']
        length = df.loc[i]['Length']
        hieght = df.loc[i]['Height']
        discharge = df.loc[i]['Discharge']
        print("\tCatchment: {}, Length: {}, Hieght: {}, Q: {}".format(catchment, length, hieght, discharge))

        sht.range('E2').value =i
        sht.range('E3').value =District
        sht.range('E4').value =latitude
        sht.range('E5').value =longitude
        sht.range('E6').value =catchment
        sht.range('E7').value =length
        sht.range('E8').value =hieght
        sht.range('E17').value =discharge

    #
    #     ws['e2'].value = i
    #     ws['e3'].value = District
    #     ws['e4'].value = latitude
    #     ws['e5'].value = longitude
    #     ws['e6'].value = catchment
    #     ws['e7'].value = length
    #     ws['e8'].value = hieght
    #     ws['e17'].value = discharge
    #
        save_path = os.path.join(output_folder, "id{}_anicut.xlsx".format(i))
        wb.save(save_path)
    #
    wb.close()
        # print(i)
        # print(df.loc[i]['Catchment'])
    # for i, row in df.iterrows():
    #     print(i, list(row))
    #     # print("Working for point: ID={}".format())
    #
    #     book = openpyxl.load_workbook(project_file_path)
    #     sheets = book.sheetnames
    #     ws = book[sheets[0]]
    #
    #     print([
    #         ws['e7'].value, ws['e8'].value, ws['e13'].value, ws['e18'].value
    #     ])


# def main():
#     input = r"F:\Watershed_Works\2102 Feb 2021\4_Anicut_Report\TestData\input_test.xlsx"
#     output = r"F:\Watershed_Works\2102 Feb 2021\4_Anicut_Report\TestData\output_test.xlsx"
#
#     book = openpyxl.load_workbook(input)
#     sheets = book.sheetnames
#     ws = book[sheets[0]]
#     print(sheets, ws)
#
#     print(ws['e5'].value)
#     ws['e5'].value = 'there'
#     print(ws['e5'].value)
#
#     # filename = ''  # This can be the same filename or a different one
#     book.save(output)

main()
